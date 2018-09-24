import pymongo
from openpyxl import load_workbook
from datetime import datetime, timedelta, timezone

from datetime import timezone

# OBSERVERA att date från MongoDB blir ett python datetime-objekt. grymt!
def utc_to_local(utc_dt):
    return utc_dt.replace(tzinfo=timezone.utc).astimezone(tz=None)


wb = load_workbook("./tomtark.xlsx")
wb.template=False
print(wb.sheetnames)

sheet = wb["Blad1"]

myclient = pymongo.MongoClient("mongodb://localhost:27017/")
mydb = myclient["mydb"]
mycol = mydb["Skarpe Nord"]

mydoc = mycol.find().sort("date")

excelRowCount = 1

sheet.cell(excelRowCount, 1).value = "Vecka"
sheet.cell(excelRowCount, 2).value = "Datum"
sheet.cell(excelRowCount, 3).value = "Tid"
sheet.cell(excelRowCount, 4).value = "Serie"
sheet.cell(excelRowCount, 5).value = "Hemmalag"
sheet.cell(excelRowCount, 6).value = "Bortalag"
sheet.cell(excelRowCount, 7).value = "Match ID"

excelRowCount = 2

for x in mydoc:
  newWeek = utc_to_local(x["date"]).isocalendar()[1]
  print(x["date"])
  date = x["date"] + timedelta(hours=1)
  print(date)
  dateString1 = date.strftime("%a, %d %b")
  dateString2 = date.strftime("%H:%M")
  sheet.cell(excelRowCount, 1).value = newWeek
  sheet.cell(excelRowCount, 2).value = dateString1
  sheet.cell(excelRowCount, 3).value = dateString2
  sheet.cell(excelRowCount, 4).value = x["series"]
  sheet.cell(excelRowCount, 5).value = x["homeTeamName"]
  sheet.cell(excelRowCount, 6).value = x["awayTeamName"]
  sheet.cell(excelRowCount, 7).value = x["gameID"]
  excelRowCount = excelRowCount + 1;

wb.save("genereratark.xlsx")
